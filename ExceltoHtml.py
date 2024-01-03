import pandas as pd
from openpyxl import load_workbook

# Load Excel file and read sheet
excel_file = 'Portfolio.xlsx'
excel_data = pd.read_excel(excel_file)
wb = load_workbook(excel_file)
sheet = wb.active

# Convert to HTML
html_table = excel_data.to_html(classes='styled-table', index=False)

# Extract cell colors and generate CSS
css_styles = ''
for row in sheet.iter_rows():
    for cell in row:
        if cell.fill.start_color.index != '00000000':
            color = cell.fill.start_color.rgb
            css_styles += f'.styled-table tr:nth-child({cell.row}):nth-child({cell.column}) td:nth-child({cell.column}) {{ background-color: {color}; }}\n'

# HTML code with added CSS style
html_with_style = f'''
<!DOCTYPE html>
<html>
<head>
    <title>Portfolio</title>
    <style>
        /* Define your CSS styles here */
        .styled-table {{
            width: 100%;
            border-collapse: collapse;
            font-family: Arial, sans-serif;
        }}
        
        .styled-table th, .styled-table td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }}
        
        .styled-table th {{
            background-color: #f2f2f2;
        }}
        
        /* Generated cell color styles */
        {css_styles}
    </style>
</head>
<body>
    <h1>Portfolio</h1>
    {html_table}
</body>
</html>
'''

# Save HTML to a file
with open('docs/index.html', 'w') as file:
    file.write(html_with_style)
